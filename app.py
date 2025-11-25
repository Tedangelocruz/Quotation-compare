import os
import sqlite3
import json
from flask import Flask, request, jsonify, send_from_directory
from pypdf import PdfReader
import re

app = Flask(__name__, static_folder='static')
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# DB Init
def init_db():
    conn = sqlite3.connect('quotations.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS quotations 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT, upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS items 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, quotation_id INTEGER, 
                  supplier_name TEXT, product_name TEXT, sku TEXT, quantity REAL, 
                  unit_price REAL, total_price REAL)''')
    conn.commit()
    conn.close()

init_db()

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

def parse_number(num_str):
    # Remove common currency symbols and whitespace
    clean_str = num_str.upper().replace('$', '').replace('€', '').replace('S/', '').replace('USD', '').replace('EUR', '').strip()
    
    # Remove internal spaces (e.g. "1 200.00")
    clean_str = clean_str.replace(' ', '')
    
    if not clean_str: return None

    # Check for Spanish format (1.234,56) vs English (1,234.56)
    # Heuristic: 
    # If ',' is the last separator and it's after a '.', it's Spanish.
    # If ',' is the only separator, it's Spanish decimal.
    # If '.' is the only separator, it's English decimal.
    
    try:
        if ',' in clean_str and '.' in clean_str:
            last_comma = clean_str.rfind(',')
            last_dot = clean_str.rfind('.')
            if last_comma > last_dot: # 1.234,56
                clean_str = clean_str.replace('.', '').replace(',', '.')
            else: # 1,234.56
                clean_str = clean_str.replace(',', '')
        elif ',' in clean_str: 
            # Ambiguous: 123,456 (English int) or 123,45 (Spanish float)?
            # If it has 3 digits after comma, likely English thousands.
            # But in prices, 3 decimals is rare. 2 is common.
            parts = clean_str.split(',')
            if len(parts[-1]) == 2: # Likely decimal
                clean_str = clean_str.replace(',', '.')
            elif len(parts[-1]) == 3: # Likely thousands
                clean_str = clean_str.replace(',', '')
            else: # Fallback to decimal
                clean_str = clean_str.replace(',', '.')
        
        return float(clean_str)
    except ValueError:
        return None

import google.generativeai as genai

def extract_with_llm(text, api_key):
    try:
        print(f"Configuring Gemini with API key: {api_key[:10]}...")
        genai.configure(api_key=api_key)
        
        # Try to list available models first
        try:
            print("Fetching available models...")
            for m in genai.list_models():
                if 'generateContent' in m.supported_generation_methods:
                    print(f"  - {m.name}")
        except Exception as e:
            print(f"Could not list models: {e}")
        
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        prompt = """
        You are a parser that converts PDF price quotations into structured line items.
        You MUST always return a JSON object with a top-level items array.
        Each element in items is an object with:

        supplier_name (string)
        product_name (string)
        product_id (string or null)
        quantity (number or null)
        unit_price (number or null)
        tax_amount (number or null)
        transport_cost (number or null)
        total_price (number or null)

        Rules:
        If the PDF is messy or unclear, make your best reasonable guess.
        If some value is missing or not numeric, use null instead of skipping the whole item.
        Never return an empty items array. If you can only find one rough line item, return that.
        Do not include explanations or comments – output ONLY valid JSON.
        
        Here is the text content of the PDF:
        """ + text

        print("Calling Gemini API...")
        response = model.generate_content(prompt)
        print(f"Gemini response received: {len(response.text)} characters")
        
        # Clean up response to ensure it's just JSON
        content = response.text
        print(f"Raw response (first 200 chars): {content[:200]}")
        
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            content = content.split("```")[1].split("```")[0]
        
        print(f"Cleaned JSON (first 200 chars): {content[:200]}")
        data = json.loads(content)
        
        # Handle both formats: direct array or object with items key
        if isinstance(data, list):
            items = data
        else:
            items = data.get('items', [])
        
        print(f"Successfully extracted {len(items)} items via LLM")
        return items
    except Exception as e:
        print(f"LLM Extraction failed: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def extract_items_from_text(text):
    # Fallback Heuristic Extraction
    if not text or len(text.strip()) < 10:
        print("Text is empty or too short")
        return [] 

    print(f"Extracting from text (first 500 chars): {text[:500]}")
    items = []
    lines = text.split('\n')
    
    supplier_name = "Unknown Supplier"
    skip_words = ["FACTURA", "QUOTATION", "PRESUPUESTO", "FECHA", "DATE", "PAGINA", "PAGE", "NIT", "RUC"]
    for line in lines[:15]:
        l = line.strip().upper()
        if l and not any(w in l for w in skip_words) and len(l) > 3:
            supplier_name = line.strip()
            break

    # Enhanced header/metadata detection
    header_keywords = [
        "DOCUMENTO", "RNC:", "CLIENTE:", "VENDEDOR:", "CONDICION:", "VENCE:",
        "HORA:", "FECHA:", "REFERENCIA:", "TELEFONO", "TEL:", "LOCAL",
        "REPARTO", "DIAS", "PÁGINA", "PAGE", "CANT.", "PRECIO", "DESC.",
        "ITBIS", "IMPORTE", "DESCRIPCIÓN", "DESCRIPCION"
    ]
    
    # Strategy 1: Look for lines that appear to be product data
    for line in lines:
        line = line.strip()
        if not line or len(line) < 10: continue
        
        # Skip headers and metadata
        upper_line = line.upper()
        is_header = False
        for keyword in header_keywords:
            if keyword in upper_line and len(line) < 60:  # Headers are usually shorter
                is_header = True
                break
        
        if is_header:
            continue
        
        # Skip lines that are clearly not product data
        if line.startswith("Página") or line.startswith("Cliente:") or line.startswith("Vendedor:"):
            continue
            
        parts = line.split()
        if len(parts) < 3:  # Need at least product code, description, and some numbers
            continue
        
        # Scan from RIGHT to LEFT to find trailing numbers (the data columns)
        # Stop when we hit text (the description)
        trailing_numbers = []
        text_parts = []
        found_text_after_numbers = False
        
        for i in range(len(parts) - 1, -1, -1):
            p = parts[i]
            val = parse_number(p)
            
            if val is not None and val > 0 and val < 1000000:
                if not found_text_after_numbers:
                    # Still in the number columns at the end
                    trailing_numbers.insert(0, val)
                else:
                    # We already found text, so this number is part of description
                    # (like "EXIST: 78")
                    text_parts.insert(0, p)
            else:
                # Hit text
                found_text_after_numbers = True
                text_parts.insert(0, p)
        
        # We need at least 2 trailing numbers to be a valid product line
        if len(trailing_numbers) >= 2:
            # Extract product ID (first alphanumeric part)
            product_id = None
            description_parts = []
            
            for part in text_parts:
                if product_id is None and any(c.isdigit() for c in part) and any(c.isalpha() for c in part):
                    product_id = part
                else:
                    description_parts.append(part)
            
            description = " ".join(description_parts[:20]) if description_parts else " ".join(text_parts[:20])
            
            # Now parse the trailing numbers
            # Expected order: Qty, Price, Discount, Tax, Total
            qty = 1.0
            unit_price = 0.0
            total_price = 0.0
            tax = None
            
            if len(trailing_numbers) >= 5:
                # Full set: Qty, Price, Discount, Tax, Total
                qty = trailing_numbers[0]
                unit_price = trailing_numbers[1]
                # trailing_numbers[2] is discount (0.00 usually)
                tax = trailing_numbers[3]
                total_price = trailing_numbers[4]
            elif len(trailing_numbers) >= 4:
                # Qty, Price, Tax, Total (discount might be 0 and hidden)
                qty = trailing_numbers[0]
                unit_price = trailing_numbers[1]
                tax = trailing_numbers[2]
                total_price = trailing_numbers[3]
            elif len(trailing_numbers) >= 3:
                # Qty, Price, Total
                qty = trailing_numbers[0]
                unit_price = trailing_numbers[1]
                total_price = trailing_numbers[2]
            elif len(trailing_numbers) == 2:
                # Price, Total
                unit_price = trailing_numbers[0]
                total_price = trailing_numbers[1]
                qty = 1.0
            
            # Sanity check on qty
            if qty > 10000:
                # Unreasonable, probably wrong interpretation
                qty = 1.0
            
            items.append({
                "supplier_name": supplier_name,
                "product_name": description.strip(),
                "product_id": product_id,
                "quantity": qty,
                "unit_price": unit_price,
                "tax_amount": tax,
                "transport_cost": None,
                "total_price": total_price
            })

    print(f"Strategy 1 found {len(items)} items")

    # If we still have no items, be more aggressive but filter better
    if len(items) == 0:
        print("Strategy 1 failed, trying aggressive fallback...")
        for line in lines:
            line = line.strip()
            if not line or len(line) < 10: continue
            
            # Stricter header filtering
            upper_line = line.upper()
            if any(kw in upper_line for kw in header_keywords):
                continue
            
            parts = line.split()
            nums = []
            text_parts = []
            for p in parts:
                v = parse_number(p)
                if v is not None and v > 0 and v < 1000000:  # Reasonable range
                    nums.append(v)
                else:
                    text_parts.append(p)
            
            if len(nums) >= 1 and text_parts:
                price = nums[-1]
                desc = " ".join(text_parts[:10])
                
                items.append({
                    "supplier_name": supplier_name,
                    "product_name": desc.strip(),
                    "product_id": None,
                    "quantity": 1,
                    "unit_price": price,
                    "tax_amount": None,
                    "transport_cost": None,
                    "total_price": price
                })

    print(f"Total items extracted: {len(items)}")
    return items

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    api_key = request.form.get('api_key')
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)
    
    try:
        reader = PdfReader(filepath)
        text = ""
        for page in reader.pages:
            extracted = page.extract_text()
            if extracted: text += extracted + "\n"
        
        if not text.strip():
             return jsonify({'error': 'No text found in PDF. This might be a scanned image (OCR required).'}), 400

        items = []
        if api_key and len(api_key) > 10:
            print("Using LLM for extraction...")
            items = extract_with_llm(text, api_key)
            if not items:
                print("LLM returned no items, falling back to heuristic.")
                items = extract_items_from_text(text)
        else:
            items = extract_items_from_text(text)
        
        if not items:
             return jsonify({'error': 'Could not extract any items.'}), 400
        
        # Save to DB
        conn = sqlite3.connect('quotations.db')
        c = conn.cursor()
        c.execute("INSERT INTO quotations (filename) VALUES (?)", (file.filename,))
        quotation_id = c.lastrowid
        
        saved_items = []
        for item in items:
            # Handle potential nulls from LLM
            qty = item.get('quantity') or 0
            price = item.get('unit_price') or 0
            total = item.get('total_price') or (qty * price)
            product_id = item.get('product_id', None)
            
            c.execute("INSERT INTO items (quotation_id, supplier_name, product_name, sku, quantity, unit_price, total_price) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (quotation_id, item.get('supplier_name', 'Unknown'), item.get('product_name', 'Unknown'), 
                       product_id, qty, price, total))
            
            item['id'] = c.lastrowid
            saved_items.append(item)
        conn.commit()
        conn.close()
        
        return jsonify({'quotation_id': quotation_id, 'items': saved_items})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/items', methods=['GET'])
def get_items():
    conn = sqlite3.connect('quotations.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM items")
    rows = c.fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])

@app.route('/api/items/<int:id>', methods=['PUT'])
def update_item(id):
    data = request.json
    conn = sqlite3.connect('quotations.db')
    c = conn.cursor()
    c.execute("UPDATE items SET product_name=?, quantity=?, unit_price=?, total_price=? WHERE id=?",
              (data['product_name'], data['quantity'], data['unit_price'], data['total_price'], id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/export', methods=['GET'])
def export_csv():
    import csv
    from io import StringIO
    from flask import Response
    
    conn = sqlite3.connect('quotations.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get the most recent quotation ID
    c.execute("SELECT id FROM quotations ORDER BY upload_date DESC LIMIT 1")
    latest_quotation = c.fetchone()
    
    if latest_quotation:
        # Only export items from the latest quotation
        c.execute("SELECT * FROM items WHERE quotation_id = ?", (latest_quotation['id'],))
        rows = c.fetchall()
    else:
        rows = []
    
    conn.close()
    
    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['ID', 'Quotation ID', 'Supplier', 'Product Name', 'SKU', 'Quantity', 'Unit Price', 'Total Price'])
    
    # Write data
    for row in rows:
        writer.writerow([row['id'], row['quotation_id'], row['supplier_name'], 
                        row['product_name'], row['sku'] or '', row['quantity'], 
                        row['unit_price'], row['total_price']])
    
    output.seek(0)
    return Response(output.getvalue(), mimetype='text/csv', 
                   headers={"Content-Disposition": "attachment;filename=quotations_export.csv"})

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import Response
    
    conn = sqlite3.connect('quotations.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get the most recent quotation ID
    c.execute("SELECT id FROM quotations ORDER BY upload_date DESC LIMIT 1")
    latest_quotation = c.fetchone()
    
    if latest_quotation:
        # Only export items from the latest quotation
        c.execute("SELECT * FROM items WHERE quotation_id = ?", (latest_quotation['id'],))
        rows = c.fetchall()
    else:
        rows = []
    
    conn.close()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations"
    
    # Define headers
    headers = ['ID', 'Quotation ID', 'Supplier', 'Product Name', 'Product ID', 'Quantity', 'Unit Price', 'Total Price']
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row in rows:
        # Get product_id from the items - check if sku column exists, otherwise use None
        product_id = row['sku'] if 'sku' in row.keys() else None
        
        ws.append([
            row['id'], 
            row['quotation_id'], 
            row['supplier_name'],
            row['product_name'], 
            product_id or '',
            row['quantity'], 
            row['unit_price'], 
            row['total_price']
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment;filename=quotations_export.xlsx"}
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
